import base64
import io
import logging

from odoo import fields, models
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:
    openpyxl = None
    _logger.warning("openpyxl not installed — Excel import for primes will not work.")


class HrPrimeImportWizard(models.TransientModel):
    _name = "hr.prime.import.wizard"
    _description = "Import de primes depuis Excel"

    campaign_id = fields.Many2one("hr.prime.campaign", required=True)
    file = fields.Binary(string="Fichier Excel")
    filename = fields.Char()
    template_file = fields.Binary(string="Modèle", readonly=True)
    template_filename = fields.Char()

    def action_download_template(self):
        """Generate an Excel template with all active employees."""
        if not openpyxl:
            raise UserError("La librairie openpyxl n'est pas installée sur le serveur.")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Primes"

        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        headers = ["Matricule", "Nom", "Département", "Montant"]
        for col, title in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=title)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Fill with active employees
        employees = self.env["hr.employee"].search(
            [("company_id", "=", self.campaign_id.company_id.id)],
            order="matricule, name",
        )
        for row_idx, emp in enumerate(employees, start=2):
            ws.cell(row=row_idx, column=1, value=emp.matricule or "")
            ws.cell(row=row_idx, column=2, value=emp.name or "")
            ws.cell(row=row_idx, column=3, value=emp.department_id.name or "")
            ws.cell(row=row_idx, column=4, value=0)

        # Column widths
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 25
        ws.column_dimensions["D"].width = 15

        buf = io.BytesIO()
        wb.save(buf)
        self.template_file = base64.b64encode(buf.getvalue())
        self.template_filename = "modele_primes.xlsx"

        # Reopen wizard to show download link
        return {
            "type": "ir.actions.act_window",
            "res_model": self._name,
            "res_id": self.id,
            "view_mode": "form",
            "target": "new",
            "context": self.env.context,
        }

    def action_import(self):
        if not openpyxl:
            raise UserError("La librairie openpyxl n'est pas installée sur le serveur.")

        if not self.file:
            raise UserError("Veuillez sélectionner un fichier Excel.")

        data = base64.b64decode(self.file)
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(min_row=2, values_only=True))
        if not rows:
            raise UserError("Le fichier est vide.")

        Employee = self.env["hr.employee"]
        existing = self.campaign_id.line_ids.mapped("employee_id.id")
        lines_vals = []
        errors = []
        skipped = 0

        for row_idx, row in enumerate(rows, start=2):
            if len(row) < 4:
                continue

            # Column A = matricule, D = montant
            matricule = str(row[0]).strip() if row[0] else ""
            amount = row[3]

            if not matricule:
                continue

            try:
                amount = float(amount or 0)
            except (ValueError, TypeError):
                errors.append(f"Ligne {row_idx}: montant invalide '{row[3]}'")
                continue

            if amount <= 0:
                skipped += 1
                continue

            employee = Employee.search(
                [("matricule", "=", matricule), ("company_id", "=", self.campaign_id.company_id.id)],
                limit=1,
            )
            if not employee:
                errors.append(f"Ligne {row_idx}: matricule '{matricule}' introuvable")
                continue

            if employee.id in existing:
                errors.append(f"Ligne {row_idx}: {employee.name} déjà dans la campagne")
                continue

            lines_vals.append({
                "campaign_id": self.campaign_id.id,
                "employee_id": employee.id,
                "amount": amount,
            })

        if errors:
            raise UserError("Erreurs lors de l'import :\n" + "\n".join(errors))

        if not lines_vals:
            msg = "Aucune ligne valide trouvée."
            if skipped:
                msg += f" ({skipped} ligne(s) ignorée(s) avec montant à 0)"
            raise UserError(msg)

        self.env["hr.prime.line"].create(lines_vals)

        return {
            "type": "ir.actions.client",
            "tag": "display_notification",
            "params": {
                "title": "Import réussi",
                "message": f"{len(lines_vals)} ligne(s) importée(s).",
                "type": "success",
                "sticky": False,
            },
        }
